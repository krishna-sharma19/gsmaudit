import sqlite3

from openpyxl import load_workbook
#import DB_manager

class DB_manager():
    
    def __init__(self):
        self.conn = sqlite3.connect(':memory:')
        self.cur = self.conn.cursor()

    def execute_create_insert(self,query):
        self.cur.execute(query)

    def execute_select(self,query):
        return self.cur.execute(query)

db = DB_manager()
db.execute_create_insert('''CREATE TABLE audit_master
                (lac, key, sector, site_name, 
               cell_id, ncc, bcc, bcch, scrambling_code)''')        

master_fields = {
        'lac' : '', 
        'key' : '', 
        'sector' : '', 
        'site_name' : '', 
        'cell_id' : '', 
        'ncc' : '', 
        'bcc' : '', 
        'bcch' : '', 
        'scrambling_code' : ''
}

file_structure_mapping= {
    'MEID-213': {
        'header': 5,
        'lac':'C',
        'key': 'D',
        'sector': 'E',
        'site_name': 'F',
        'cell_id': 'G',
        'ncc': 'H',
        'bcc': 'I',
        'bcch': 'T',
    },
    'MEID-214': {
        'header': 5,
        'lac' : 'C',
        'key': 'D',
        'sector': 'E',
        'site_name': 'F',
        'cell_id': 'G',
        'ncc': 'H',
        'bcc': 'I',
        'bcch': 'T',
    },
    'External': {
        'header': 5,
        'lac' : 'G',
        'key': 'D',
        'sector': 'E',
        'site_name': 'F',
        'cell_id': 'H',
        'ncc': 'K',
        'bcc': 'L',
        'bcch': 'J',
    },

}





class Importer():
    """
    This is SqlLite data importor factory.
    Given axlsx file and sheet names import data 
    from excel into database.
    """
    def check_for_errors(self):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        print('selecting records')
        
        for row in self.get_rows_from_sheet(self.filename, self.sheetname, file_structure_mapping['External'] ):
            search_string = row['cell_id'] + row['lac'] #2131 lac in external
            search_query = "select cell_id||lac as unikey from audit_master where unikey='"
            search_query = search_query + search_string +"'"
            #import pdb
            #pdb.set_trace()
            #print(row['cell_id']+'_'+row['lac'])
            #print(search_query)
            #print('Found-')
            #print([i.value for i in db.execute_select(search_query)])
            #print(row)

    def get_rows_from_sheet(self, filename, sheetname, field_template):
        wb = load_workbook(filename=filename)
        sheet = wb[sheetname]
        
        for row in sheet.iter_rows(row_offset=field_template['header']):
            true_row = dict()
            for k,v in field_template.items():
                if k is not 'header' and row[ord(v)-65].value is not None:
                    true_row[k]=row[ord(v)-65].value
            #print(true_row)        
            if len(true_row)!=0: 
                for k,v in master_fields.items():
                    if not true_row.get(k):
                        true_row[k] = ''
                yield(true_row)            
        
    def load_table_data(self, lookup=None, filename=None, sheetname=None):
        wb = load_workbook(filename=filename)
        sheet = wb[sheetname]
        input_table = {}
        for row in self.get_rows_from_sheet(self.filename, self.sheetname,lookup):
            query = """INSERT INTO audit_master VALUES('{lac}','{key}','{sector}','{site_name}','{cell_id}','{ncc}','{bcc}','{bcch}','{scrambling_code}');""".format(**row)
            db.execute_create_insert(query)
            print(row['cell_id']+'_'+row['lac'])
            #print('This row inserted successfully !^^^^')

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        global db

        if self.sheetname is 'GGsmCell':
            self.load_table_data(
                    lookup=file_structure_mapping['MEID-213'], 
                    filename=self.filename,
                    sheetname=self.sheetname)
        elif self.sheetname is 'GExternalGsmCell':
            self.check_for_errors()
              
